import pandas as pd
import openpyxl

# Load workbook
wb = openpyxl.load_workbook('qx.xlsx')
ws = wb['bedah']

# Baca ke dataframe untuk memudahkan
headers = [cell.value for cell in ws[1]]
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)
df = pd.DataFrame(data, columns=headers)

# Fungsi untuk menambahkan kontradiksi
def add_contradiction(surat, ayat, text):
    mask = (df['surat'] == surat) & (df['ayat'] == ayat)
    idx = df[mask].index
    if len(idx) > 0:
        current = df.loc[idx[0], 'CONTRADICTIONS']
        if pd.isna(current) or str(current).strip() == '':
            df.loc[idx[0], 'CONTRADICTIONS'] = text
        else:
            # Sudah ada, jangan timpa
            pass

# === SURAT 12 (Yusuf) ===
add_contradiction(12, 1, 'Kontradiksi Kejelasan Al-Quran: Ayat ini berisi huruf muqatta\'ah "Alif Lam Ra" tanpa makna jelas — bertentangan dengan klaim Quran sebagai "kitab yang menjelaskan segala sesuatu" (16:89) dan "bahasa Arab yang nyata" (16:103); 29 surat mengandung huruf-huruf ini tanpa pernah dijelaskan maknanya oleh Quran itu sendiri.')
add_contradiction(12, 3, 'Kontradiksi dengan 6:14 dan Klaim Kenabian: Ayat ini menyatakan Muhammad "sebelumnya termasuk orang yang tidak mengetahui" kisah Yusuf — bertentangan dengan klaim kenabian yang menyiratkan wahyu memberikan pengetahuan komprehensif, dan juga bertentangan dengan 42:52 yang menyatakan Muhammad tidak mengetahui apa itu kitab dan iman sebelum wahyu.')
add_contradiction(12, 38, 'Kontradiksi Kronologi Agama: Yusuf mengaku mengikuti "agama nenek moyangku: Ibrahim, Ishak dan Yakub" — ini anakronistik karena Taurat (hukum Yahudi) belum ada pada zaman Yusuf; konsep "agama" yang terikat pada nenek moyang tertentu bertentangan dengan klaim Islam sebagai agama universal.')
add_contradiction(12, 76, 'Kontradiksi Keadilan dan Meritokrasi: Ayat ini menyatakan "Kami angkat derajat orang yang Kami kehendaki" — bertentangan dengan prinsip keadilan proporsional yang menghargai usaha manusia; jika derajat hanya bergantung pada kehendak arbitrer Allah, maka sistem moral kehilangan prediktabilitas dan keadilan.')
add_contradiction(12, 83, 'Kontradiksi Narasi Kronologi: Ayat ini mengulang kalimat dari ayat 18 ("hanya dirimu sendiri yang memandang baik urusan yang buruk itu") dalam konteks yang berbeda (kehilangan Benyamin vs Yusuf) — menunjukkan penggabungan naratif yang rancu dan inkonsisten secara internal.')
add_contradiction(12, 100, 'Kontradiksi dengan Larangan Sujud: Yusuf menerima sujud dari kedua orang tuanya — bertentangan dengan larangan sujud kepada makhluk dalam Islam (13:15 menyatakan sujud hanya untuk Allah); jika sujud ini hanya hormat, mengapa Iblis dihukum karena menolak sujud kepada Adam?')

# === SURAT 13 (Ar-Ra\'d) ===
add_contradiction(13, 2, 'Kontradiksi Lokasi Allah: Ayat ini menyatakan Allah "bersemayam di atas Arsy" — bertentangan dengan 2:115 yang menyatakan "ke mana pun kamu menghadap, di situlah wajah Allah" (imanensi total); dua konsep ini—transsendensi di Arsy vs kehadiran universal—sulit dikompromikan secara logis.')
add_contradiction(13, 11, 'Kontradiksi Takdir dan Kebebasan: Ayat ini menyatakan "Allah tidak mengubah keadaan suatu kaum sebelum mereka mengubah keadaan diri mereka sendiri" — bertentangan dengan 9:51 ("tidak akan menimpa kami kecuali apa yang telah ditetapkan Allah untuk kami") dan 35:8 (Allah menyesatkan siapa yang Dia kehendaki); jika Allah mengendalikan segala sesuatu, maka perubahan diri kaum tidak relevan.')
add_contradiction(13, 27, 'Kontradiksi Penyebab Kesesatan: Ayat ini menyatakan "Allah menyesatkan siapa yang Dia kehendaki" — bertentangan dengan 2:256 ("tidak ada paksaan dalam agama") dan 6:108 (manusia menyesatkan diri sendiri); jika Allah yang sengaja menyesatkan, maka konsep keadilan dan pertanggungjawaban individu runtuh.')
add_contradiction(13, 31, 'Kontradiksi Internal Teks: Teks standar berbunyi "a-fa-lam yay\'asi" (tidakkah orang beriman berputus asa?), sedangkan Ibn Abbas—sahabat Nabi dan mufasir utama—membaca "yatabayyan" (tidakkah orang beriman melihat dengan jelas?). Dua bacaan kanonik yang valid menghasilkan makna yang berlawanan total, menunjukkan ketidakstabilan teks pada masa formatif.')
add_contradiction(13, 38, 'Kontradiksi dengan 13:7: Ayat ini menyatakan "tidak ada hak bagi seorang rasul mendatangkan bukti (mukjizat) melainkan dengan izin Allah" — namun ayat 13:7 sendiri menyatakan Muhammad hanyalah "seorang pemberi peringatan" yang tidak bisa mendatangkan tanda; inkonsistensi ini menunjukkan tekanan teologis berbeda pada fase Makkah vs Madinah.')

# === SURAT 14 (Ibrahim) ===
add_contradiction(14, 4, 'Kontradiksi dengan 2:256 (Tidak Ada Paksaan): Ayat ini menyatakan "Allah menyesatkan siapa yang Dia kehendaki" — bertentangan langsung dengan prinsip "tidak ada paksaan dalam agama" (2:256); jika Allah sengaja menyesatkan sebagian manusia, maka pada praktiknya terdapat paksaan takdir yang mematikan kebebasan memilih.')
add_contradiction(14, 19, 'Kontradiksi Urutan Penciptaan: Ayat ini menyatakan Allah menciptakan "langit dan bumi dengan hak" — namun tidak menjelaskan urutannya; bertentangan dengan 79:27-30 (langit dibangun SEBELUM bumi) dan 41:9-12 (bumi diciptakan SEBELUM langit dalam 2 hari, baru langit dalam 4 hari).')
add_contradiction(14, 22, 'Kontradiksi Peran Iblis: Setan berkata "aku tidak membenarkan perbuatanmu mempersekutukanku" — menyiratkan setan tidak setuju dengan penyembahan berhala; bertentangan dengan 15:39-40 di mana Iblis dengan bangga mengaku akan menyesatkan manusia, dan 38:82-83 di mana Iblis meminta izin Allah untuk menyesatkan.')
add_contradiction(14, 27, 'Kontradiksi dengan 2:256: Ayat ini menyatakan "Allah menyesatkan orang-orang yang zalim" — kembali menegaskan Allah sebagai agen aktif penyesatan; bertentangan dengan konsep keadilan ilahi yang mengharuskan manusia bertanggung jawab atas pilihannya sendiri.')
add_contradiction(14, 52, 'Kontradiksi Kejelasan Al-Quran: Ayat ini menyatakan Al-Quran adalah "penjelasan (yang sempurna) bagi manusia" — bertentangan dengan 3:7 yang mengakui adanya ayat mutasyabihat yang maknanya hanya diketahui Allah, dan 67:3 yang menyatakan tidak ada yang inkonsisten kecuali mata manusia yang buta.')

# === SURAT 15 (Al-Hijr) ===
add_contradiction(15, 9, 'Kontradiksi dengan Fakta Historis: Ayat ini mengklaim "Kamilah yang menurunkan Al-Quran dan pasti Kami pula yang memeliharanya" — bertentangan dengan temuan manuskrip Sana\'a, Birmingham, dan Topkapi yang memuat puluhan ribu varian teks, rasm, dan bacaan yang berbeda; jika Allah benar-benar memelihara teks, mengapa ada korupsi dan variasi yang begitu masif?')
add_contradiction(15, 26, 'Kontradiksi Cara Penciptaan Manusia: Ayat ini menyatakan manusia diciptakan dari "tanah liat kering dari lumpur hitam" — bertentangan dengan 16:4 (dari mani), 96:2 (dari darah clot/alaq), 21:30 (dari air), dan 32:7-8 (dari tanah kemudian dari mani); inkonsistensi internal tentang bahan penciptaan manusia.')
add_contradiction(15, 27, 'Kontradiksi Penciptaan Jin: Ayat ini menyatakan jin diciptakan "dari api yang sangat panas" (samaq) — bertentangan dengan 55:15 yang menyatakan jin diciptakan dari "api tanpa asap" (samum); dua jenis api berbeda untuk makhluk yang sama menunjukkan inkonsistensi naratif.')
add_contradiction(15, 34, 'Kontradiksi dengan 2:35: Ayat ini menyatakan Iblis diusir dari surga — bertentangan dengan 2:35-36 yang menggambarkan Adam (bukan Iblis) di surga, dan Iblis diusir SETELAH menolak sujud kepada Adam; kontradiksi ini menunjukkan dua versi penciptaan yang digabungkan secara tidak mulus.')
add_contradiction(15, 85, 'Kontradiksi Urutan Penciptaan: Ayat ini menyatakan Allah tidak menciptakan langit dan bumi "melainkan dengan kebenaran" — namun bertentangan dengan 79:27-33 dan 41:9-12 tentang urutan penciptaan yang berbeda; jika penciptaan itu "benar," mengapa ada deskripsi yang saling bertentangan?')

# === SURAT 16 (An-Nahl) ===
add_contradiction(16, 4, 'Kontradiksi Cara Penciptaan Manusia: Ayat ini menyatakan manusia diciptakan dari "mani" — bertentangan dengan 15:26 (tanah liat kering), 96:2 (darah clot), dan 32:7-8 (tanah lalu mani); inkonsistensi internal tentang bahan asal penciptaan manusia.')
add_contradiction(16, 40, 'Kontradiksi Cara Penciptaan: Ayat ini menyatakan Allah menciptakan dengan berkata "Kun fayakun" (jadilah, maka jadilah) secara instan — bertentangan dengan 7:54, 10:3, 41:9-12, dan 50:38 yang menyatakan Allah menciptakan langit dan bumi dalam enam hari secara bertahap.')
add_contradiction(16, 43, 'Kontradiksi Siapa yang Bisa Jadi Rasul: Ayat ini menyatakan "Kami tidak mengutus sebelum engkau melainkan orang laki-laki yang Kami beri wahyu" — bertentangan dengan 22:75 yang menyatakan "Allah memilih utusan dari malaikat dan dari manusia"; malaikat jelas bisa menjadi rasul/utusan bagi manusia.')

# === SURAT 17 (Al-Isra\') ===
add_contradiction(17, 12, 'Kontradiksi Kejelasan Al-Quran: Ayat ini menyatakan "segala sesuatu telah Kami terangkan dengan jelas" — bertentangan dengan 3:7 yang mengakui adanya ayat mutasyabihat (yang hanya Allah mengetahui maknanya) dan 67:3 yang menyatakan tidak ada inkonsistensi kecuali mata yang buta.')
add_contradiction(17, 23, 'Kontradiksi dengan 9:23: Ayat ini memerintahkan berbakti kepada orang tua — bertentangan dengan 9:23 yang melarang taat kepada orang tua jika mereka memanggil kepada kesyirikan, dan 29:8 serta 31:15 yang menyatakan jangan taat kepada orang tua jika mereka menyuruh berbuat syirik; kontradiksi antara taat mutlak vs taat bersyarat.')
add_contradiction(17, 46, 'Kontradiksi dengan 2:256: Ayat ini menyatakan Allah "menutup hati dan menyumbat telinga" orang kafir — bertentangan dengan prinsip "tidak ada paksaan dalam agama" (2:256); jika Allah secara aktif menghalangi pemahaman, maka kekafiran bukanlah pilihan bebas melainkan hasil intervensi ilahi yang memaksa.')
add_contradiction(17, 50, 'Kontradiksi Identitas Iblis: Ayat ini menyatakan Iblis "adalah dari golongan jin" — bertentangan dengan 2:34 dan 38:71-76 yang menyatakan Iblis ada di antara para malaikat; jin dan malaikat adalah dua makhluk yang berbeda dalam teologi Islam, sehingga ini adalah kontradiksi identitas fundamental.')
add_contradiction(17, 70, 'Kontradiksi dengan 17:97: Ayat ini menyatakan Allah "memuliakan anak cucu Adam" — bertentangan dengan 17:97 yang menggambarkan orang kafir dikumpulkan "dengan wajah tersungkur, dalam keadaan buta, bisu, dan tuli"; jika manusia dimuliakan, mengapa ada yang dihina secara fisik dan sensorik?')
add_contradiction(17, 97, 'Kontradiksi dengan 75:22-23: Ayat ini menyatakan orang kafir datang pada hari Kiamat "dengan wajah tersungkur, buta, bisu, dan tuli" — bertentangan dengan 75:22-23 yang menyatakan pada hari itu "wajah-wajah berseri-seri" (untuk orang beriman); namun deskripsi sensorik yang dihapuskan secara total untuk orang kafir bertentangan dengan konsep hisab (perhitungan) yang memerlukan indra untuk memberi kesaksian.')
add_contradiction(17, 101, 'Kontradiksi Jumlah Mukjizat Musa: Ayat ini menyatakan Musa diberi "sembilan mukjizat yang nyata" — bertentangan dengan 7:133 yang menyebutkan sembilan wabah/azab (bukan mukjizat), dan 27:12 yang menyebutkan satu tanda (tangan bersinar); inkonsistensi tentang apa yang dihitung sebagai mukjizat/tanda.')
add_contradiction(17, 106, 'Kontradiksi dengan 2:185: Ayat ini menyatakan Al-Quran diturunkan "berangsur-angsur" (bertahap) — bertentangan dengan 2:185 dan 97:1 yang menyatakan Al-Quran diturunkan sekaligus pada Lailatul Qadar; dua mode penurunan (bertahap vs sekaligus) tidak dapat didamaikan tanpa kontradiksi.')

# === SURAT 18 (Al-Kahf) ===
add_contradiction(18, 9, 'Kontradiksi dengan Sumber Historis: Ayat ini memulai kisah Ashabul Kahfi dengan tanda tanya retoris — namun kisah ini secara historis merupakan plagiarisme dari fabel Kristen Suriah "The Seven Sleepers of Ephesus" (abad ke-5 M); kontradiksi klaim "wahyu" vs adaptasi dari dongeng lokal yang sudah ada sebelum Islam.')
add_contradiction(18, 25, 'Kontradiksi Waktu Kiamat: Ayat ini menyatakan Ashabul Kahfi tinggal di gua "tiga ratus tahun dan ditambah sembilan tahun" — bertentangan dengan 20:103 yang menyatakan "sepersekian sehari" (hitungan singkat), dan 32:5 (satu hari dengan Allah = 1.000 tahun menurut perhitungan manusia); inkonsistensi skala waktu eskatologis.')
add_contradiction(18, 27, 'Kontradiksi dengan 2:106: Ayat ini menyatakan "tidak ada yang dapat mengubah kalimat-kalimat (kitab) Allah" — bertentangan dengan 2:106 (naskh: ayat diganti dengan ayat lain), 16:101 (penggantian ayat), dan 13:39 (Allah menghapus dan menetapkan apa yang Dia kehendaki).')
add_contradiction(18, 29, 'Kontradiksi dengan 2:7: Ayat ini menyatakan "barang siapa menghendaki beriman, hendaklah dia beriman" — bertentangan dengan 2:7 (Allah mengunci hati orang kafir), 4:88 (Allah membiarkan mereka sesat), dan 32:32 (Allah menyesatkan siapa yang Dia kehendaki); jika Allah mengontrol kehendak manusia, maka pilihan beriman adalah ilusi.')
add_contradiction(18, 47, 'Kontradiksi dengan 16:15: Ayat ini menyatakan gunung-gunung akan "diperjalankan" (dihancurkan/dipindahkan) pada hari Kiamat — bertentangan dengan 16:15 yang menyatakan gunung "ditancapkan" (diyakinkan tidak bergerak) agar bumi tidak goncang; jika gunung permanen, mengapa bisa diperjalankan?')
add_contradiction(18, 50, 'Kontradiksi dengan 2:34: Ayat ini menyatakan Iblis "dari golongan jin" yang mendurhakai — bertentangan dengan 2:34 dan 38:71-76 yang menempatkan Iblis di antara para malaikat; jin dan malaikat adalah dua kategori makhluk yang berbeda dalam teks, sehingga ini kontradiksi klasifikasi.')
add_contradiction(18, 60, 'Kontradiksi dengan Sifat Nabi: Kisah Musa mencari Khidir yang kemudian melakukan pembunuhan anak dan perusakan perahu — bertentangan dengan sifat nabi yang harusnya merupakan teladan moral; jika seorang "hamba Allah" (Khidir) dapat membunuh tanpa proses hukum, ini merusak konsep keadilan proporsional.')
add_contradiction(18, 86, 'Kontradiksi Sains: Ayat ini menggambarkan Zulkarnain melihat matahari terbenam "di dalam laut yang berlumpur hitam" — ini adalah kosmologi datar geosentris yang bertentangan dengan fakta astronomis bahwa matahari jauh lebih besar dari bumi dan tidak terbenam di laut manapun.')

# === SURAT 19 (Maryam) ===
add_contradiction(19, 1, 'Kontradiksi Kejelasan Al-Quran: Ayat ini berisi huruf muqatta\'ah "Kaf Ha Ya \'Ain Sad" tanpa makna jelas — bertentangan dengan 16:89 (Quran menjelaskan segala sesuatu) dan 12:1 (ayat-ayat yang jelas); keberadaan huruf-huruf ini di 29 surat tanpa penjelasan adalah kontradiksi internal.')
add_contradiction(19, 10, 'Kontradiksi dengan 3:41: Ayat ini menyatakan Zakaria tidak bisa bercakap-cakap "selama tiga malam" — bertentangan dengan 3:41 yang menyatakan "tiga hari" (siang); perbedaan durasi dan waktu (malam vs siang) untuk peristiwa yang sama menunjukkan inkonsistensi naratif.')
add_contradiction(19, 28, 'Kontradiksi Historis: Ayat ini menyebut Maryam (Ibu Yesus, abad 1 M) sebagai "saudara perempuan Harun" — mengasosiasikannya dengan Miryam saudara Musa dan Harun dari era Eksodus (sekitar 1200 SM); ini adalah anakronisme kronologis dan genealogis yang fatal, menunjukkan ketidaktahuan penulis teks akan sejarah.')
add_contradiction(19, 34, 'Kontradiksi dengan 4:157: Ayat ini menyatakan "Itulah Isa putra Maryam, (yang mengatakan) perkataan yang benar" — namun 4:157 menyatakan orang Yahudi "tidak membunuhnya dan tidak menyalibnya" melainkan "disamakan baginya orang lain"; jika Yesus benar-benar wafat (seperti 19:33), maka 4:157 yang menyangkal kematian Yesus adalah kontradiksi langsung.')
add_contradiction(19, 71, 'Kontradiksi dengan 39:61: Ayat ini menyatakan "tidak ada seorang pun di antara kamu yang tidak mendatanginya (neraka)" — bertentangan dengan 39:61 (Allah menyelamatkan orang bertakwa dari neraka), 21:101-102 (orang yang telah Kami janjikan akan dijauhkan dari neraka), dan 52:18 (mereka tidak merasakan panas neraka); jika semua harus mendatangi neraka, maka janji keselamatan adalah palsu.')

# Update workbook dari dataframe
for idx, row in df.iterrows():
    for col_idx, col_name in enumerate(headers, 1):
        ws.cell(row=idx+2, column=col_idx, value=row[col_name])

wb.save('qx.xlsx')
print('Done updating qx.xlsx')
print('Total contradictions added:')
for s in range(12, 20):
    subset = df[(df['surat'] == s)]
    nama = str(subset['Surat IDN'].iloc[0]).replace('\n', ' ') if len(subset) > 0 else '???'
    total = len(subset)
    ada = subset['CONTRADICTIONS'].notna().sum()
    print(f'  Surat {s} ({nama}): {total} ayat, {ada} sudah ada contradictions')
